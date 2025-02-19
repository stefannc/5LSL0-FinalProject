# 5LSL0 Final Assignment main code for spectrum based model

## Import arbitrary libraries
import numpy as np
import matplotlib.pyplot as plt
import os
import openpyxl
import sys

## Import tensorflow and keras
import tensorflow as tf
from tensorflow import keras
from tensorflow.keras.layers import Input, Dense, Conv1D, Conv2D, MaxPooling1D, LeakyReLU, AvgPool2D, UpSampling2D, ReLU, MaxPooling2D, Reshape, Flatten, Activation
from tensorflow.keras.losses import MSE, categorical_crossentropy
from tensorflow.keras.models import Model, Sequential
from tensorflow.keras.optimizers import Adam
from sklearn.metrics import accuracy_score
from keras.models import load_model
import bisect

## Import selfmade modules
from generator_spectrum import DataGenerator
from model_spectrum import deep_cnn
sys.path.insert(0, 'Data')
import createNoise
from datasplitter import datasplitter

def getDivisors(x):
    divisors = []
    i = 1
    for i in range(1,x):
        if(x%i == 0): 
            divisors.append(i)
        else: 
            continue
        i+=1
    return divisors

def find_le(a, x):
    'Find rightmost value less than or equal to x'
    i = bisect.bisect_right(a, x)
    if i:
        return a[i-1]
    raise ValueError

## Labels
labels=("yes", "no", "up", "down", "left", "right", "on", "off", "stop", "go", "silence", "unknown")
num_classes = len(labels)
training_batch_size = 128
# test_batch_size = 28
final_test_batch_size = 6


## normalization options
normalization = {"subtract_mean": True,
                 "epsilon": 1e-8,           # prevent division by 0
                 "normalize": "95",        # "none", "var", "std", "95", "97"
                 "type": "sample"}          # "sample", "class"

## evaluation metrics
metrics=['acc']

## create silence
USE_SILENCE = True
N_SILENCE = 2000
if USE_SILENCE:
    createNoise.create(N_SILENCE)
    
## create lists
datasplitter()

## paths
data_path = 'Data/train/audio/'
data_lists = 'Data/train/lists/'
final_test_path = 'Data/test/audio/'

## Create generator
traingen1 = DataGenerator(data_path=data_path,
                         data_listing=data_lists+'train_set.txt',
                         batch_size=training_batch_size, 
                         dims_in=(16000,1),
                         dims_out=(99,161),#(99,161),#(13,32), 
                         labels=labels)

validgen1 = DataGenerator(data_path=data_path,
                         data_listing=data_lists+'validation_set.txt',
                         batch_size=training_batch_size, 
                         dims_in=(16000,1),
                         dims_out=(99,161),#(99,161),#(13,32),
                         labels=labels)

## Create test model
shape = (99,161,1) #(99,161,1) 
model = deep_cnn(shape, num_classes)

#optimizer = Adam(lr = 0.0001)
model.compile(optimizer='Adam', loss = categorical_crossentropy, metrics = metrics)

# Train model on dataset
num_lines_train = sum(1 for line in open(data_lists+'train_set.txt'))
num_lines_valid = sum(1 for line in open(data_lists+'validation_set.txt'))

history = model.fit_generator(generator=traingen1,
                              epochs = 12,
                              verbose = 1,
                              validation_data=validgen1)

# summarize history for acc
plt.plot(history.history['acc'])
plt.plot(history.history['val_acc'])
plt.title('model accuracy')
plt.ylabel('accuracy')
plt.xlabel('epoch')
plt.legend(['train', 'valid'], loc='upper left')
plt.savefig('results/spectrum/model_spectrum_norm_acc.pdf')
plt.close()

# summarize history for loss
plt.plot(history.history['loss'])
plt.plot(history.history['val_loss'])
plt.title('model loss')
plt.ylabel('loss')
plt.xlabel('epoch')
plt.legend(['train', 'valid'], loc='upper left')
plt.savefig('results/spectrum/model_spectrum_norm_loss.pdf')

# save model
model.save('results/spectrum/model_spectrum_norm.h5')

## test data
from generator_spectrum_test import DataGenerator

with open(data_lists+'test_set.txt') as f:
    content = f.readlines()

divisors_test = getDivisors(len(content))
test_batch_size = find_le(divisors_test, 100)

testgen = DataGenerator(data_path=data_path,
                         data_listing=data_lists+'test_set.txt',
                         batch_size=test_batch_size, 
                         dims_in=(16000,1),
                         dims_out=(99,161),#(99,161),#(13,32),
                         labels=labels)

y_pred_proba = model.predict_generator(generator=testgen, 
                                     verbose=1)

y_pred = np.argmax(y_pred_proba, axis=1)



y_true = []
# Added modulo to test batch size to make sure len(y_true) == len(y_pred)
for lbl in range(len(content)):
    label = content[lbl].split('/')[0]
    if label in labels:
        y_true.append(labels.index(label))
    else:
        y_true.append(labels.index('unknown'))
        
acc_score = accuracy_score(y_true, y_pred)    
print(acc_score)

## test final data for submission
from generator_spectrum_finaltest import DataGenerator

y_files = os.listdir(final_test_path)

finaltestgen = DataGenerator(data_path=final_test_path,
                             dataset = y_files,
                             batch_size=final_test_batch_size,
                             dims_in=(16000,1),
                             dims_out=(99,161),#(99,161),#
                             labels=labels)

y_pred_proba = model.predict_generator(generator = finaltestgen,
                                     verbose=1)
y_pred = np.argmax(y_pred_proba, axis=1)

# Save computed probabilities so that they can be used with the other models
np.savetxt('results/spectrum/probabilities_spectrum_norm.csv', y_pred_proba, delimiter=',')

# Save the final outcome
wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value='fname,label'
for y in range(len(y_pred)):
    sheet.cell(row=y+2, column=1).value=y_files[y]+','+labels[y_pred[y]]

wb.save('results/spectrum/submission_spectrum_norm.xlsx')