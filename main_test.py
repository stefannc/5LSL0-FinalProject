"""
=============================================================================
    Eindhoven University of Technology
==============================================================================

    Source Name   : main.py
                    Main ML file for the 5LSL0 project
    Author        : Bart van Erp
    Date          : 06/06/2019

==============================================================================
"""

## Import arbitrary libraries
import numpy as np
import matplotlib.pyplot as plt
import os
import openpyxl

## Import tensorflow and keras
import tensorflow as tf
from tensorflow import keras
from tensorflow.keras.layers import Input, Dense, Conv1D, Conv2D, MaxPooling1D, LeakyReLU, AvgPool2D, UpSampling2D, ReLU, MaxPooling2D, Reshape, Flatten, Activation
from tensorflow.keras.losses import MSE, categorical_crossentropy
from tensorflow.keras.models import Model, Sequential
from tensorflow.keras.optimizers import Adam
from sklearn.metrics import accuracy_score

## Import selfmade moduels
from Processing.generator_mfcc import DataGenerator
from model import deep_cnn


## Labels
labels=("yes", "no", "up", "down", "left", "right", "on", "off", "stop", "go", "silence", "unknown")
num_classes = len(labels)
batch_size = 128


## normalization options
normalization = {"subtract_mean": True,
                 "epsilon": 1e-8,           # prevent division by 0
                 "normalize": "95",        # "none", "var", "std", "95", "97"
                 "type": "sample"}          # "sample", "class"

## evaluation metrics
metrics=['acc']

data_path = 'Data/train/audio/'
data_lists = 'Data/train/lists/'
final_test_path = 'Data/test/audio/' # unzipped train and test data

## Create generator
traingen1 = iter(DataGenerator(data_path=data_path,
                         data_listing=data_lists+'train_set.txt',
                         batch_size=128, 
                         dims_in=(16000,1),
                         dims_out=(13,32),
                         labels=labels))

validgen1 = iter(DataGenerator(data_path=data_path,
                         data_listing=data_lists+'validation_set.txt',
                         batch_size=128, 
                         dims_in=(16000,1),
                         dims_out=(13,32),
                         labels=labels))

## Create test model
shape = (13,32,1)
model = deep_cnn(shape, num_classes)

#optimizer = Adam(lr = 0.0001)
model.compile(optimizer='Adam', loss = categorical_crossentropy, metrics = metrics)

# Train model on dataset
num_lines_train = sum(1 for line in open(data_lists+'train_set.txt'))
num_lines_valid = sum(1 for line in open(data_lists+'validation_set.txt'))

history = model.fit_generator(generator=traingen1,
                              steps_per_epoch= int(np.ceil(num_lines_train/batch_size)),
                              epochs = 15,
                              verbose = 1,
                              validation_data=validgen1,
                              validation_steps = int(np.ceil(num_lines_valid/batch_size)))


plt.plot(history.history['acc'])
plt.plot(history.history['val_acc'])
plt.title('model accuracy')
plt.ylabel('accuracy')
plt.xlabel('epoch')
plt.legend(['train', 'valid'], loc='upper left')
plt.show()
# summarize history for loss
plt.plot(history.history['loss'])
plt.plot(history.history['val_loss'])
plt.title('model loss')
plt.ylabel('loss')
plt.xlabel('epoch')
plt.legend(['train', 'valid'], loc='upper left')
plt.show()

# save model
model.save('model.h5')

## test data
from Processing.generator_mfcc_test import DataGenerator

testgen = iter(DataGenerator(data_path=data_path,
                         data_listing=data_lists+'test_set.txt',
                         batch_size=128, 
                         dims_in=(16000,1),
                         dims_out=(13,32),
                         labels=labels))

num_lines_test = sum(1 for line in open(data_lists+'test_set.txt'))

y_pred_proba = model.predict_generator(generator=testgen, 
                                     steps=int(np.ceil(len(num_lines_test)/batch_size)), 
                                     verbose=1)

y_pred = np.argmax(y_pred_proba, axis=1)

with open(data_lists+'test_set.txt') as f:
    content = f.readlines()

y_true = []
for lbl in range(len(content)):
    label = content[lbl].split('/')[0]
    if label in labels:
        y_true.append(labels.index(label))
    else:
        y_true.append(labels.index('unknown'))
        
acc_score = accuracy_score(y_true, y_pred)    
print(acc_score)



## test final data for submission


#y_files = os.listdir(final_test_path)
#y_pred_proba = model.predict_generator(generator = finaltestgen, 
#                                     steps = int(np.ceil(len(y_files)/batch_size)), 
#                                     verbose=1)
#y_pred = np.argmax(y_pred_proba, axis=1)



#wb = openpyxl.Workbook()
#sheet = wb.active
#sheet.cell(row=1, column=1).value='fname,label'
#for y in range(len(y_pred)):
#    sheet.cell(row=y+2, column=1).value=y_files[y]+','+labels[y_pred[y]]

#wb.save('example.xlsx')
